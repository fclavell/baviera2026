#!/usr/bin/env python3
"""
sync_excel.py — Sincronitza recarregues.json amb el full "Recàrregues" de Baviera 2026.xlsx
Ús: python3 sync_excel.py [recarregues.json] [Baviera 2026.xlsx]
"""

import json, sys, os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Cal instal·lar openpyxl: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
JSON_FILE = Path(sys.argv[1]) if len(sys.argv) > 1 else BASE / "app" / "recarregues.json"
EXCEL_FILE = Path(sys.argv[2]) if len(sys.argv) > 2 else BASE / "Baviera 2026.xlsx"

if not JSON_FILE.exists():
    print(f"No s'ha trobat: {JSON_FILE}")
    sys.exit(1)
if not EXCEL_FILE.exists():
    print(f"No s'ha trobat: {EXCEL_FILE}")
    sys.exit(1)

# ── Llegir JSON ────────────────────────────────────────────────────────────
with open(JSON_FILE, encoding="utf-8") as f:
    data = json.load(f)

recarregues = data.get("recarregues", [])
km_log      = data.get("km", [])
print(f"Llegint {JSON_FILE.name}: {len(recarregues)} recàrregues, {len(km_log)} entrades km")

# ── Obrir Excel ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE)

# ── Full RECÀRREGUES ───────────────────────────────────────────────────────
sheet_name = "Recàrregues"
if sheet_name not in wb.sheetnames:
    # Prova variacions d'accent
    for n in wb.sheetnames:
        if "rec" in n.lower():
            sheet_name = n
            break

ws = wb[sheet_name]
print(f"Full trobat: '{sheet_name}'")

# Estructura del full (de l'Excel original):
# Col 0 = descripció/dia  Col 1 = data  Col 2 = res  Col 3 = Proveïdor
# Col 4 = Preu  Col 5 = kWh  Col 6 = Preu kWh
# Les files de dades comencen aprox a fila 7 (index 6)
# Cada dia ocupa 3 files; fila 6 = Dia 1 (08-10), etc.

# Mapeig: dia 1→fila 7, dia 2→fila 10, ... (3 files per dia)
# Però l'Excel original té les files de resum als dies
# Escriurem a partir de la fila 7 (1-indexed) netejant primer les files de dades

HEADER_ROWS = 6   # les 6 primeres files són capçalera/resum
ROWS_PER_DAY = 3
FIRST_DATA_ROW = HEADER_ROWS + 1  # fila 7 (1-indexed)
TOTAL_DAYS = 14

# Neteja les files de dades existents (deixem capçaleres intactes)
last_data_row = FIRST_DATA_ROW + TOTAL_DAYS * ROWS_PER_DAY
for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=last_data_row, min_col=4, max_col=7):
    for cell in row:
        cell.value = None

# Escriure les recàrregues agrupades per dia
# Cada recàrrega va a la primera cel·la disponible del seu dia
day_slots = {}  # dia_num → llista de recàrregues
for r in recarregues:
    dia = r.get("dia", 1)
    if dia not in day_slots:
        day_slots[dia] = []
    day_slots[dia].append(r)

written = 0
for dia, recs in day_slots.items():
    if dia < 1 or dia > TOTAL_DAYS:
        continue
    base_row = FIRST_DATA_ROW + (dia - 1) * ROWS_PER_DAY
    for slot, r in enumerate(recs[:ROWS_PER_DAY]):  # màx 3 per dia
        row_idx = base_row + slot
        ws.cell(row=row_idx, column=4).value = r.get("xarxa", "")
        kwh  = r.get("kwh",  0) or 0
        preu = r.get("preu", 0) or 0
        ws.cell(row=row_idx, column=5).value = round(preu, 2) if preu else None
        ws.cell(row=row_idx, column=6).value = round(kwh, 2)  if kwh  else None
        ws.cell(row=row_idx, column=7).value = round(preu/kwh, 4) if kwh and preu else None
        written += 1

print(f"Recàrregues escrites al full: {written}")

# ── Resum per xarxa (columnes 12-15, files 2-5) ───────────────────────────
xarxes_order = ["Casa", "Tesla", "Ionity", "Altres"]
totals = {x: {"kwh": 0.0, "preu": 0.0} for x in xarxes_order}
for r in recarregues:
    x = r.get("xarxa", "Altres")
    if x not in totals:
        x = "Altres"
    totals[x]["kwh"]  += r.get("kwh",  0) or 0
    totals[x]["preu"] += r.get("preu", 0) or 0

# Files 2-5 (1-indexed), columnes 12=kWh, 13=Total €, 14=€/kWh
for i, xarxa in enumerate(xarxes_order):
    row_idx = 2 + i
    t = totals[xarxa]
    ws.cell(row=row_idx, column=12).value = round(t["kwh"],  2) if t["kwh"]  else 0
    ws.cell(row=row_idx, column=13).value = round(t["preu"], 2) if t["preu"] else 0
    ws.cell(row=row_idx, column=14).value = round(t["preu"]/t["kwh"], 4) if t["kwh"] else 0

# Total (fila 6 del resum, columna 12-14)
total_kwh  = sum(t["kwh"]  for t in totals.values())
total_preu = sum(t["preu"] for t in totals.values())
ws.cell(row=6, column=12).value = round(total_kwh,  2) if total_kwh  else 0
ws.cell(row=6, column=13).value = round(total_preu, 2) if total_preu else 0
ws.cell(row=6, column=14).value = round(total_preu/total_kwh, 4) if total_kwh else 0

# ── Desar ──────────────────────────────────────────────────────────────────
ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
out  = EXCEL_FILE.parent / f"Baviera 2026_sync_{ts}.xlsx"
wb.save(out)
print(f"\n✅ Excel guardat a: {out.name}")

print("\n── Resum per xarxa ──────────────────────────")
print(f"{'Xarxa':<10} {'kWh':>8} {'Total €':>10} {'€/kWh':>8}")
print("-" * 42)
for xarxa in xarxes_order:
    t = totals[xarxa]
    ekwh = t["preu"]/t["kwh"] if t["kwh"] else 0
    print(f"{xarxa:<10} {t['kwh']:>8.2f} {t['preu']:>10.2f} {ekwh:>8.4f}")
print("-" * 42)
print(f"{'TOTAL':<10} {total_kwh:>8.2f} {total_preu:>10.2f} {total_preu/total_kwh:>8.4f}" if total_kwh else f"{'TOTAL':<10} {'0':>8} {'0':>10} {'—':>8}")
